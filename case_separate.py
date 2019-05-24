import re
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

def get_list(case_text):
    """
    获取编号列表和文本列表
    """
    case_text = case_text.lstrip()
    try:
        first_num_str = re.search(r'\d{4,5}\s+',case_text).group() #\s匹配空格，group() 同group（0）就是匹配正则表达式整体结果
    except Exception as e:
        print(e)
    try:
        first_num = int(first_num_str)
    except Exception as e:
        print(e)

    n_list = [first_num]  # 序号列表

    # 1. 把每组数据标号拿出来
    n_set = re.findall(r'\n\n\d{4,5}\s+', case_text)
    for i in n_set:
        n_list.append(int(i))

    # 2. 把每组数据的内容拿出来
    c_list = re.split(r'\n\n\d{4,5}\s+', case_text)

    if len(first_num_str) == 4:
        c_list[0] = c_list[0][5:]
    else:
        c_list[0] = c_list[0][6:]
    return n_list,c_list


def get_proposal(case_list):
    test_proposal_list = []
    stress_test_list = []
    HT_test_list = []
    for i in range(0,len(case_list)):
        try:
            a = re.search(r'Test-Proposal(.*)\n', case_list[i], re.I).group(1)
            if 'Y' in a:
                test_proposal_list.append('Y')
            elif "N" in a:
                test_proposal_list.append('N')
            else:
                test_proposal_list.append('NA')
        except Exception as e:
            print(e)
            test_proposal_list.append('NA')
        try:
            a = re.search(r'Stress-Test(.*)\n', case_list[i], re.I).group(1)
            if 'Y' in a:
                stress_test_list.append('Y')
            elif "N" in a:
                stress_test_list.append('N')
            else:
                stress_test_list.append('NA')
        except Exception as e:
            print(e)
            stress_test_list.append('NA')
        try:
            a = re.search(r'HW-Test(.*)', case_list[i], re.I).group(1)
            if 'Y' in a:
                HT_test_list.append('Y')
            elif "N" in a:
                HT_test_list.append('N')
            else:
                HT_test_list.append('NA')
        except Exception as e:
            print(e)
            HT_test_list.append('NA')
    return test_proposal_list, stress_test_list, HT_test_list


def set_style_write(sheet, write_row, write_column, write_value):
    sheet.cell(row=write_row, column=write_column).value = write_value
    sheet.cell(row=write_row, column=write_column).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=write_row, column=write_column).font = Font(u'Arial', size = 10, bold=True)


def write_excel(excel_name,case_text):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = '修改点'
    # 设置第一行行高和ABC列宽
    sheet1.row_dimensions[1].height = 15
    sheet1.column_dimensions['A'].width = 15
    sheet1.column_dimensions['B'].width = 90
    sheet1.column_dimensions['C'].width = 15
    sheet1.column_dimensions['D'].width = 15
    sheet1.column_dimensions['E'].width = 15
    sheet1.column_dimensions['F'].width = 15

    n_list, c_list = get_list(case_text)
    test_proposal_list, stress_test_list, HT_test_list = get_proposal(c_list)

    row0 = ['修改点编号', '修改点描述', '执行人', 'Test-Proposal', 'Stress-Test', 'HW-Test']
    for k in range(0, len(row0)):
        sheet1.cell(row=1, column=k+1).value = row0[k]
        # 设置第一行关键字水平垂直居中
        sheet1.cell(row=1, column=k+1).alignment = Alignment(horizontal='center', vertical='center')
        # 设置字体
        sheet1.cell(row=1, column=k+1).font = Font(u'Arial', size = 10, bold=True)
        # 填充背景色
        sheet1.cell(row=1, column=k+1).fill = PatternFill("solid", fgColor="969696")
    # 写入A列修改点编号
    for i in range(0, len(n_list)):
        set_style_write(sheet1, i+2, 1, n_list[i])
    # 写入B列修改点详情
    for j in range(0, len(c_list)):
        sheet1.cell(row=j+2, column=2).value = c_list[j]
        # 设置修改点内容字体
        sheet1.cell(row=j+2, column=2).alignment = Alignment(wrap_text=True)
        # 设置修改点内容自动换行
        sheet1.cell(row=j+2, column=2).font = Font(u'Arial', size = 10, bold=True)
    # 写入第D列Stress-Test
    for j in range(0, len(test_proposal_list)):
        set_style_write(sheet1, j+2, 4, test_proposal_list[j])
    # 写入第E列Test-Proposal
    for j in range(0, len(stress_test_list)):
        set_style_write(sheet1, j+2, 5, stress_test_list[j])
    # 写入第F列HW-Test
    for j in range(0, len(HT_test_list)):
        set_style_write(sheet1, j+2, 6, HT_test_list[j])

    wb.save(excel_name)